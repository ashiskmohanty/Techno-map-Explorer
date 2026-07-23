"""
ETL for PS Process Explorer dashboard.

Reads the three source Excel workbooks, normalises them into a single
`data.json` consumed by the front-end dashboard (index.html).

Sources
-------
1. Latest PE Objects - Unused code.xlsx   -> ABAP custom objects
2. SAP PSPE - BWIP Object Analysis.xlsx   -> BW (BW-IP) custom objects
3. BPML V1.0.xlsx                         -> Business Process Master List (L1/L2/L3)

Run:  python build_data.py
"""
from __future__ import annotations

import glob
import json
import os
import re
from datetime import date, datetime

HERE = os.path.dirname(os.path.abspath(__file__))

ABAP_FILE = "Latest PE Objects - Unused code.xlsx"
BW_FILE = "SAP PSPE - BWIP Object Analysis.xlsx"
BPML_FILE = "BPML V1.0.xlsx"

CUSTOM_PREFIXES = ("Z", "Y", "/CPD/", "/1CPMB/")
PACKAGES = ("ZPS_PROJ_EXEC", "Z_PROF_SERVICES", "ZCPM")


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def s(v):
    """Safe string."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def is_custom(name: str) -> bool:
    n = (name or "").upper().lstrip()
    return n.startswith(CUSTOM_PREFIXES)


def norm_pa(v: str) -> str:
    v = s(v)
    if not v or v in ("#REF!", "#N/A", "0"):
        return ""
    return v


# --------------------------------------------------------------------------- #
# Readers
# --------------------------------------------------------------------------- #
def read_xlsx(path):
    """Return {sheet_name: [ [cells...], ... ]} for an OOXML .xlsx file."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append([c for c in r])
        out[ws.title] = rows
    wb.close()
    return out


def read_legacy_xls(path):
    """Return {sheet_name: [ [cells...], ... ]} for a legacy OLE2 .xls file."""
    import xlrd

    book = xlrd.open_workbook(path)
    out = {}
    for sh in book.sheets():
        rows = []
        for rx in range(sh.nrows):
            row = []
            for cx in range(sh.ncols):
                cell = sh.cell(rx, cx)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate.xldate_as_datetime(val, book.datemode)
                    except Exception:
                        pass
                row.append(val)
            rows.append(row)
        out[sh.name] = rows
    return out


def load_workbook(name):
    path = os.path.join(HERE, name)
    if not os.path.exists(path):
        print(f"  ! missing {name}")
        return {}
    with open(path, "rb") as fh:
        head = fh.read(2)
    try:
        if head == b"PK":
            return read_xlsx(path)
        return read_legacy_xls(path)
    except Exception as e:
        # last resort: try the other reader
        try:
            return read_legacy_xls(path) if head == b"PK" else read_xlsx(path)
        except Exception:
            print(f"  ! failed to read {name}: {e}")
            return {}


def header_index(rows):
    """Find the first row that looks like a header, return (idx, {name:col})."""
    for i, r in enumerate(rows[:8]):
        vals = [s(c).lower() for c in r]
        filled = [v for v in vals if v]
        if len(filled) >= 2:
            return i, {s(c): j for j, c in enumerate(r) if s(c)}
    return 0, {}


def rows_as_dicts(rows):
    idx, hdr = header_index(rows)
    keys = list(hdr.keys())
    cols = list(hdr.values())
    out = []
    for r in rows[idx + 1:]:
        if not any(s(c) for c in r):
            continue
        d = {}
        for k, c in zip(keys, cols):
            d[k] = r[c] if c < len(r) else None
        out.append(d)
    return out


# --------------------------------------------------------------------------- #
# Domain extraction
# --------------------------------------------------------------------------- #
def classify_abap(name: str) -> str:
    n = (name or "").upper()
    if n.startswith("TABLEFRAME_"):
        return "Table Maintenance"
    if "_CL_" in n or n.startswith(("ZCL_", "YCL_", "ZCX_")):
        return "Class"
    if n.startswith(("ZIF_", "YIF_")):
        return "Interface"
    return "Function Module"


def build():
    print("Reading workbooks...")
    abap_wb = load_workbook(ABAP_FILE)
    bw_wb = load_workbook(BW_FILE)
    bpml_wb = load_workbook(BPML_FILE)

    objects = []          # unified custom-object list
    process_areas = {}    # name -> aggregation

    def ensure_pa(name):
        name = name or "Unassigned"
        if name not in process_areas:
            process_areas[name] = {
                "name": name,
                "counts": {},
            }
        return process_areas[name]

    def add_obj(o):
        objects.append(o)
        pa = ensure_pa(o["process"])
        pa["counts"][o["category"]] = pa["counts"].get(o["category"], 0) + 1

    # ---- ABAP objects ----------------------------------------------------- #
    for sheet, rows in abap_wb.items():
        for d in rows_as_dicts(rows):
            name = s(d.get("Object name") or d.get("ABAP Object name")
                     or d.get("Object Name"))
            if not name:
                continue
            add_obj({
                "name": name,
                "domain": "ABAP",
                "category": classify_abap(name),
                "process": norm_pa(d.get("Process area")) or "Unassigned",
                "package": s(d.get("Development class / package")),
                "author": s(d.get("Author")),
                "created": "",
                "description": "",
                "validity": "",
                "technical": "",
                "source": f"ABAP:{sheet}",
            })

    # ---- BW: BEx Queries -------------------------------------------------- #
    for d in rows_as_dicts(bw_wb.get("BEx Queries", [])):
        name = s(d.get("Query"))
        if not name:
            continue
        add_obj({
            "name": name,
            "domain": "BW",
            "category": "BEx Query",
            "process": norm_pa(d.get("Process Area")) or "Unassigned",
            "package": "",
            "author": s(d.get("Responsible")),
            "created": s(d.get("Last Execution Date")),
            "description": s(d.get("Long description")),
            "validity": s(d.get("Validity?")),
            "technical": s(d.get("Technical Details")),
            "source": "BW:BEx Queries",
        })

    # ---- BW: Planning Sequences ------------------------------------------ #
    for d in rows_as_dicts(bw_wb.get("Planning Sequences", [])):
        name = s(d.get("Planning Sequence"))
        if not name:
            continue
        add_obj({
            "name": name,
            "domain": "BW",
            "category": "Planning Sequence",
            "process": norm_pa(d.get("Process Area")) or "Unassigned",
            "package": "",
            "author": s(d.get("Used by")),
            "created": "",
            "description": s(d.get("Description")),
            "validity": "",
            "technical": "",
            "source": "BW:Planning Sequences",
        })

    # ---- BW: Planning Functions ------------------------------------------ #
    for d in rows_as_dicts(bw_wb.get("Planning Functions", [])):
        name = s(d.get("Planning Function"))
        if not name:
            continue
        add_obj({
            "name": name,
            "domain": "BW",
            "category": "Planning Function",
            "process": norm_pa(d.get("Process Area")) or "Unassigned",
            "package": "",
            "author": "",
            "created": s(d.get("Last Used")),
            "description": s(d.get("Description")),
            "validity": s(d.get("Validity?")),
            "technical": s(d.get("Technical details")),
            "source": "BW:Planning Functions",
        })

    # ---- BW: Infoproviders ----------------------------------------------- #
    for d in rows_as_dicts(bw_wb.get("Infoproviders", [])):
        name = s(d.get("InfoProvider"))
        if not name:
            continue
        add_obj({
            "name": name,
            "domain": "BW",
            "category": "InfoProvider",
            "process": norm_pa(d.get("Process Area")) or "Unassigned",
            "package": "",
            "author": "",
            "created": "",
            "description": s(d.get("Long description")),
            "validity": s(d.get("Validity?")),
            "technical": "",
            "source": "BW:Infoproviders",
        })

    # ---- BW: Aggregation Levels ------------------------------------------ #
    for d in rows_as_dicts(bw_wb.get("Aggregation Levels", [])):
        name = s(d.get("Aggregation Level"))
        if not name:
            continue
        add_obj({
            "name": name,
            "domain": "BW",
            "category": "Aggregation Level",
            "process": norm_pa(d.get("Process Area")) or "Unassigned",
            "package": "",
            "author": "",
            "created": "",
            "description": s(d.get("Description")),
            "validity": s(d.get("Valid/Delete")),
            "technical": s(d.get("Info provider")),
            "planseq": s(d.get("Planning sequence")),
            "source": "BW:Aggregation Levels",
        })

    # ---- BW: Filters ------------------------------------------------------ #
    for d in rows_as_dicts(bw_wb.get("Filters", [])):
        name = s(d.get("Filters"))
        if not name:
            continue
        add_obj({
            "name": name,
            "domain": "BW",
            "category": "Filter",
            "process": "Unassigned",
            "package": "",
            "author": "",
            "created": "",
            "description": "",
            "validity": s(d.get("Valid/Delete")),
            "technical": s(d.get("Level")),
            "planseq": s(d.get("Planning sequence")),
            "planfunc": s(d.get("Planning Function")),
            "source": "BW:Filters",
        })

    # ---- BW: Info Objects ------------------------------------------------- #
    for d in rows_as_dicts(bw_wb.get("Info Objects", [])):
        name = s(d.get("InfoObject"))
        if not name:
            continue
        add_obj({
            "name": name,
            "domain": "BW",
            "category": "InfoObject",
            "process": "Unassigned",
            "package": "",
            "author": "",
            "created": "",
            "description": s(d.get("Description")),
            "validity": s(d.get("Validity?")),
            "technical": s(d.get("Technical Details")),
            "source": "BW:Info Objects",
        })

    # keep only custom (Z/Y/...) objects for the custom-object views
    for o in objects:
        o["custom"] = is_custom(o["name"])

    # ---- BPML hierarchy --------------------------------------------------- #
    bpml = parse_bpml(bpml_wb)

    # attach BPML areas as process areas even if no objects yet
    for node in bpml:
        ensure_pa(node["l3"] or node["l2"] or node["l1"])

    # ---- Dependency edges (BW planseq <-> planfunc <-> filter/agg) -------- #
    edges = build_edges(objects, bw_wb)

    # ---- Assemble --------------------------------------------------------- #
    pa_list = []
    for name, pa in sorted(process_areas.items()):
        total = sum(pa["counts"].values())
        pa_list.append({
            "name": name,
            "total": total,
            "counts": pa["counts"],
            "abap": sum(v for k, v in pa["counts"].items()
                        if k in ("Function Module", "Class", "Interface",
                                 "Table Maintenance")),
            "bw": sum(v for k, v in pa["counts"].items()
                      if k in ("BEx Query", "Planning Sequence",
                               "Planning Function", "InfoProvider",
                               "Aggregation Level", "Filter", "InfoObject")),
        })

    data = {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "environment": "SAP MS1 / 122",
        "packages": list(PACKAGES),
        "source": "offline",
        "processAreas": pa_list,
        "objects": objects,
        "bpml": bpml,
        "edges": edges,
        "stats": build_stats(objects, bw_wb),
    }

    out_path = os.path.join(HERE, "data.json")
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=1, ensure_ascii=False)

    # data.js fallback so index.html also works when opened via file:// (no server)
    js_path = os.path.join(HERE, "data.js")
    with open(js_path, "w", encoding="utf-8") as fh:
        fh.write("window.__PSPE_DATA__ = ")
        json.dump(data, fh, ensure_ascii=False)
        fh.write(";\n")

    print(f"Wrote {out_path}")
    print(f"  process areas : {len(pa_list)}")
    print(f"  objects       : {len(objects)}")
    print(f"  custom objects: {sum(1 for o in objects if o['custom'])}")
    print(f"  edges         : {len(edges)}")
    print(f"  bpml nodes    : {len(bpml)}")
    return data


def parse_bpml(bpml_wb):
    """Best-effort parse of the Business Process Master List into L1/L2/L3."""
    nodes = []
    for sheet, rows in bpml_wb.items():
        idx, hdr = header_index(rows)
        lower = {k.lower(): v for k, v in hdr.items()}

        def find(*subs):
            for k, v in lower.items():
                if all(sub in k for sub in subs):
                    return v
            return None

        # try named columns first
        c_l1 = find("level 1") or find("l1")
        c_l2 = find("level 2") or find("l2")
        c_l3 = find("level 3") or find("l3")
        c_l4 = find("level 4") or find("l4")
        c_scope = find("scope")

        if c_l1 is None and c_l2 is None and c_l3 is None:
            # unstructured: capture first 4 text columns heuristically
            for r in rows[idx + 1:]:
                cells = [s(c) for c in r]
                filled = [c for c in cells if c]
                if not filled:
                    continue
                nodes.append({
                    "sheet": sheet,
                    "l1": cells[0] if len(cells) > 0 else "",
                    "l2": cells[1] if len(cells) > 1 else "",
                    "l3": cells[2] if len(cells) > 2 else "",
                    "l4": cells[3] if len(cells) > 3 else "",
                    "scope": "",
                })
            continue

        for r in rows[idx + 1:]:
            def g(c):
                return s(r[c]) if c is not None and c < len(r) else ""
            l1, l2, l3, l4 = g(c_l1), g(c_l2), g(c_l3), g(c_l4)
            if not any([l1, l2, l3, l4]):
                continue
            nodes.append({
                "sheet": sheet,
                "l1": l1, "l2": l2, "l3": l3, "l4": l4,
                "scope": g(c_scope),
            })
    return nodes


def build_edges(objects, bw_wb):
    """Dependency edges among BW objects for the dependency graph."""
    edges = []
    seen = set()

    def add(src, tgt, kind):
        if not src or not tgt:
            return
        key = (src, tgt, kind)
        if key in seen:
            return
        seen.add(key)
        edges.append({"source": src, "target": tgt, "kind": kind})

    # Aggregation Level -> Planning Sequence  &  Aggregation Level -> InfoProvider
    for d in rows_as_dicts(bw_wb.get("Aggregation Levels", [])):
        agg = s(d.get("Aggregation Level"))
        add(agg, s(d.get("Planning sequence")), "uses-planseq")
        add(agg, s(d.get("Info provider")), "on-provider")

    # Filter -> Planning Sequence / Planning Function
    for d in rows_as_dicts(bw_wb.get("Filters", [])):
        flt = s(d.get("Filters"))
        add(s(d.get("Planning sequence")), flt, "planseq-filter")
        add(s(d.get("Planning Function")), flt, "planfunc-filter")
        add(s(d.get("Planning sequence")), s(d.get("Planning Function")),
            "planseq-planfunc")

    # BEx Query -> technical (ABAP class / read routine)
    for d in rows_as_dicts(bw_wb.get("BEx Queries", [])):
        q = s(d.get("Query"))
        tech = s(d.get("Technical Details"))
        if tech and tech.upper().startswith(("Z", "Y")):
            add(q, tech, "query-abap")

    return edges


def build_stats(objects, bw_wb):
    """Summary numbers for the analytics cards & charts."""
    by_cat = {}
    by_domain = {}
    by_validity = {}
    for o in objects:
        by_cat[o["category"]] = by_cat.get(o["category"], 0) + 1
        by_domain[o["domain"]] = by_domain.get(o["domain"], 0) + 1
        v = (o.get("validity") or "Unknown").title()
        by_validity[v] = by_validity.get(v, 0) + 1
    return {
        "byCategory": by_cat,
        "byDomain": by_domain,
        "byValidity": by_validity,
        "total": len(objects),
        "custom": sum(1 for o in objects if o["custom"]),
    }


if __name__ == "__main__":
    build()
