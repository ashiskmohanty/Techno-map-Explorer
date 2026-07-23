import os, glob

folder = r"c:\MS - Surface\Agent in VScode -1\PS - Process Explorer"
files = [f for f in glob.glob(os.path.join(folder, "*.xlsx")) if not os.path.basename(f).startswith("~$")]

def dump_openpyxl(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        print("-" * 80)
        print("  SHEET:", ws.title, "max_row:", ws.max_row, "max_col:", ws.max_column)
        rows = list(ws.iter_rows(values_only=True))
        for i, r in enumerate(rows[:6]):
            print("   R%d:" % i, [("" if c is None else str(c))[:28] for c in r])
    wb.close()

def dump_pandas(path):
    import pandas as pd
    xls = pd.ExcelFile(path)
    for sheet in xls.sheet_names:
        df = xls.parse(sheet, header=None, nrows=6)
        print("-" * 80)
        print("  SHEET:", sheet, "shape(preview):", df.shape)
        for i, row in df.iterrows():
            print("   R%d:" % i, [("" if pd.isna(c) else str(c))[:28] for c in row.tolist()])

for path in files:
    print("=" * 100)
    print("FILE:", os.path.basename(path))
    with open(path, "rb") as fh:
        head = fh.read(4)
    try:
        if head[:2] == b"PK":
            dump_openpyxl(path)
        else:
            dump_pandas(path)
    except Exception as e:
        print("  ERROR primary:", e)
        try:
            dump_pandas(path)
        except Exception as e2:
            print("  ERROR pandas:", e2)
