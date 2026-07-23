"""Export the Process tiles currently shown on the dashboard Home page to Excel.

Applies the same exclusions as the UI (LUMIRA*, Obselete, NWS, DV) and writes
`Process_Tiles.xlsx` next to this script.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
HIDDEN = {"OBSELETE", "OBSOLETE", "NWS", "DV", "UPDATE DOMAIN", "AO", "IRQ"}


def hidden(name: str) -> bool:
    u = (name or "").strip().upper()
    return "LUMIRA" in u or u in HIDDEN


def main():
    with open(os.path.join(HERE, "data.json"), "r", encoding="utf-8") as fh:
        data = json.load(fh)

    areas = [p for p in data.get("processAreas", []) if not hidden(p.get("name", ""))]
    areas.sort(key=lambda p: p.get("total", 0), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Process Tiles"

    headers = ["#", "Process Area", "Total Objects", "ABAP", "BW", "Object Breakdown"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1F3B6E")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for i, p in enumerate(areas, 1):
        counts = p.get("counts", {}) or {}
        breakdown = "; ".join(f"{k}: {v}" for k, v in
                              sorted(counts.items(), key=lambda kv: -kv[1]))
        ws.append([i, p.get("name", ""), p.get("total", 0),
                   p.get("abap", 0), p.get("bw", 0), breakdown])

    total_row = ["", "TOTAL", sum(p.get("total", 0) for p in areas),
                 sum(p.get("abap", 0) for p in areas),
                 sum(p.get("bw", 0) for p in areas), f"{len(areas)} tiles"]
    ws.append(total_row)

    last = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if c.column in (2, 6) else "center",
                                    wrap_text=(c.column == 6))
    for c in ws[last]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EAF0FB")

    widths = {"A": 5, "B": 46, "C": 14, "D": 8, "E": 8, "F": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{last - 1}"

    out = os.path.join(HERE, "Process_Tiles.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(areas)} tiles)")


if __name__ == "__main__":
    main()
